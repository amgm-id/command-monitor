import csv
import io
from datetime import datetime
from typing import List, Any
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
def _fmt_value(value: Any) -> str:
    """Format nilai untuk export — datetime ditampilkan apa adanya (sudah waktu lokal)."""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value) if value is not None else ""


def export_csv(data: List[dict], fieldnames: List[str]) -> bytes:
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in data:
        formatted = {
            k: _fmt_value(v) if isinstance(v, datetime) else v
            for k, v in row.items()
        }
        writer.writerow(formatted)
    return output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def export_excel(data: List[dict], fieldnames: List[str], sheet_name: str = "Report") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, field in enumerate(fieldnames, 1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 20

    risk_colors = {
        "critical": "C0392B",
        "high": "E74C3C",
        "medium": "F39C12",
        "low": "27AE60",
    }

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            value = row_data.get(field, "")
            cell_value = _fmt_value(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if field == "risk_level" and str(value).lower() in risk_colors:
                cell.font = Font(color=risk_colors[str(value).lower()], bold=True)

        if row_idx % 2 == 0:
            for col_idx in range(1, len(fieldnames) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"
                )

    for col_idx, field in enumerate(fieldnames, 1):
        max_length = max(
            len(field),
            max((len(str(row.get(field, "") or "")) for row in data), default=0),
        )
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_pdf(data: List[dict], fieldnames: List[str], title: str = "Report") -> bytes:
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=0.5 * inch, bottomMargin=0.5 * inch)

    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=12,
        textColor=colors.HexColor("#1E3A5F"),
    )
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f"Dicetak: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} WITA",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.2 * inch))

    headers = [f.replace("_", " ").title() for f in fieldnames]
    table_data = [headers]

    for row in data:
        row_values = [_fmt_value(row.get(field, "")) for field in fieldnames]
        table_data.append(row_values)

    col_width = (landscape(A4)[0] - inch) / len(fieldnames)
    col_widths = [col_width] * len(fieldnames)

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F9FA")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#DEE2E6")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)

    doc.build(elements)
    return output.getvalue()
